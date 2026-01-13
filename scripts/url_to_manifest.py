#!/usr/bin/env python3
"""Generate DataWarp manifest from ANY NHS publication URL (ZIP/XLSX/CSV).

Usage:
    python url_to_manifest.py <publication_url> <output_manifest.yaml>

Example:
    python url_to_manifest.py \\
        "https://digital.nhs.uk/.../april-2025" \\
        manifests/gp_apr_2025.yaml
"""
import sys, re, yaml, requests, tempfile
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from collections import defaultdict
from zipfile import ZipFile

MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

MONTHS_FULL = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
               'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}

def parse_period(text):
    """Extract period from text (filenames, URLs)."""
    text = text.lower()
    
    # Pattern 1: Full month-YYYY (in URL paths): 31-october-2025, october-2025
    for month, num in MONTHS_FULL.items():
        match = re.search(rf'\d*-?{month}-(\d{{4}})', text)
        if match:
            return f"{match.group(1)}-{num:02d}"
    
    # Pattern 2: Month-YYYY (4-digit year): sep-2024 → 2024-09
    match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-_](\d{4})', text)
    if match:
        return f"{match.group(2)}-{MONTHS[match.group(1)]:02d}"
    
    # Pattern 3: Month_YY (2-digit year): apr_25 → 2025-04
    match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-_](\d{2})(?:[^0-9]|$)', text)
    if match:
        return f"20{match.group(2)}-{MONTHS[match.group(1)]:02d}"
    
    # Pattern 4: Year only: 2023 → 2023-01
    match = re.search(r'[/_](\d{4})(?:[/_]|$)', text)
    if match:
        return f"{match.group(1)}-01"
    
    return None

def extract_period_from_filename(filename: str) -> str | None:
    """Extract period from filename patterns. Returns None if unable to parse.
    
    Examples:
        q2-2526 -> 2025-Q2
        oct-2025 -> 2025-10
        2025-10 -> 2025-10
        week42-2025 -> 2025-W42
    """
    import re
    
    filename_lower = filename.lower()
    
    # Quarterly: q1-2526, q2-2425, q3-2324
    match = re.search(r'q([1-4])[_-]?(\d{2})(\d{2})', filename_lower)
    if match:
        quarter, year1, year2 = match.groups()
        # q2-2526 means Q2 of 2025-26 fiscal year -> use start year
        year = f"20{year1}"
        return f"{year}-Q{quarter}"
    
    # Monthly: oct-2025, 2025-10, m10-2025
    month_map = {
        'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
        'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
        'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
    }
    
    for month_name, month_num in month_map.items():
        if month_name in filename_lower:
            # Extract year
            year_match = re.search(r'20\d{2}', filename_lower)
            if year_match:
                return f"{year_match.group()}-{month_num}"
    
    # Numeric month: 2025-10, m10-2025
    match = re.search(r'(20\d{2})[_-](\d{1,2})', filename_lower)
    if match:
        year, month = match.groups()
        return f"{year}-{month.zfill(2)}"
    
    # Weekly: week42-2025, w42-2025
    match = re.search(r'w(?:eek)?[_-]?(\d{1,2})[_-]?(20\d{2})', filename_lower)
    if match:
        week, year = match.groups()
        return f"{year}-W{week.zfill(2)}"
    
    # Annual: 2025, fy2025
    match = re.search(r'(?:fy)?[_-]?(20\d{2})(?![_-]\d)', filename_lower)
    if match and not re.search(r'(q[1-4]|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', filename_lower):
        return match.group(1)
    
    return None

def scrape_resources(url):
    """Scrape ALL downloadable resources (ZIP/XLSX/CSV)."""
    print(f"Scraping {url}...", file=sys.stderr)

    # Add browser headers to avoid 403 Forbidden errors
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1'
    }

    response = requests.get(url, headers=headers)
    response.raise_for_status()
    
    soup = BeautifulSoup(response.content, 'html.parser')
    resources = []
    
    for link in soup.find_all('a', href=True):
        href = link['href']
        lower_href = href.lower()
        
        if lower_href.endswith(('.zip', '.xlsx', '.xls', '.xlsm', '.csv')):
            full_url = urljoin(url, href)
            file_type = lower_href.split('.')[-1].upper()
            title = link.get_text(strip=True) or Path(urlparse(full_url).path).name
            
            resources.append({
                'url': full_url,
                'type': file_type,
                'title': title
            })
    
    print(f"Found {len(resources)} files", file=sys.stderr)
    return resources

def capture_publication_context(url, resources):
    """OPTIONAL: Capture context for LLM enrichment (fails gracefully)."""
    context = {}
    
    try:
        response = requests.get(url, timeout=10)
        soup = BeautifulSoup(response.content, 'html.parser')
        
        # Page title
        title = soup.find('h1')
        if title:
            context['page_title'] = title.get_text(strip=True)
        
        # Find first XLSX for metadata
        for res in resources:
            if res['type'] in ['XLSX', 'XLS', 'XLSM']:
                try:
                    from io import BytesIO
                    import openpyxl
                    from itertools import islice
                    
                    print(f"  Capturing context from {Path(res['url']).name}...", file=sys.stderr)
                    wb_response = requests.get(res['url'], timeout=30)
                    wb = openpyxl.load_workbook(BytesIO(wb_response.content), read_only=True, data_only=True)
                    
                    # First sheet content (first 100 rows to capture full table list)
                    first_sheet = wb.worksheets[0]
                    content_lines = []
                    for row in islice(first_sheet.iter_rows(values_only=True), 100):
                        if any(row):
                            # Join only non-empty cells
                            row_text = ' '.join(str(c).strip() for c in row[:10] if c and str(c).strip())
                            if row_text:
                                content_lines.append(row_text)
                    
                    context['excel_metadata'] = [{
                        'file_url': res['url'],
                        'first_sheet': wb.sheetnames[0],
                        'content_preview': '\n'.join(content_lines)[:5000]
                    }]
                    break  # Only first XLSX
                except Exception as e:
                    print(f"  ⚠ Could not extract XLSX context: {e}", file=sys.stderr)
    except Exception as e:
        print(f"⚠ Could not capture publication context: {e}", file=sys.stderr)
    
    return context if context else None

def get_excel_sheets(url):
    """Download XLSX and return sheet names. Returns empty list if file is old XLS format."""
    import openpyxl
    from io import BytesIO
    
    try:
        response = requests.get(url)
        wb = openpyxl.load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        return wb.sheetnames
    except Exception as e:
        # openpyxl only supports XLSX, not old XLS format
        if 'not a zip file' in str(e).lower() or 'xls' in url.lower():
            print(f"  ⚠ Skipping old XLS file (not supported): {Path(urlparse(url).path).name}", file=sys.stderr)
            return []
        raise  # Re-raise other errors

def inspect_zip(url):
    """Download ZIP and list CSV/XLSX files."""
    response = requests.get(url, stream=True)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    for chunk in response.iter_content(8192):
        tmp.write(chunk)
    tmp.close()
    
    with ZipFile(tmp.name) as zf:
        files = [f for f in zf.namelist() 
                if not f.endswith('/') and f.lower().endswith(('.csv', '.xlsx', '.xls'))]
    
    Path(tmp.name).unlink()
    return files

def sanitize_name(text):
    """Create clean, SHORT, DATE-AGNOSTIC code from filename."""
    name = Path(text).stem
    name = re.sub(r'%20', '_', name)
    name = re.sub(r'[^a-z0-9_]', '_', name.lower())
    name = re.sub(r'_+', '_', name).strip('_')
    
    # CRITICAL: Strip ALL date patterns for stable table names across months
    # Month_YY: apr_25, april_2025
    name = re.sub(r'_(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)_\d{2,4}', '', name, flags=re.I)
    # Full month names: january_2025, april_25
    name = re.sub(r'_(january|february|march|april|may|june|july|august|september|october|november|december)_\d{2,4}', '', name, flags=re.I)
    # Standalone years: _2025, _2024
    name = re.sub(r'_\d{4}(?:_|$)', '_', name)
    # Month without year: _april, _jan
    name = re.sub(r'_(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)(?:_|$)', '_', name, flags=re.I)
    
    # Clean up consecutive underscores again
    name = re.sub(r'_+', '_', name).strip('_')
    
    # Keep only last 2-3 meaningful segments
    parts = [p for p in name.split('_') if p and len(p) > 2]
    if len(parts) > 3:
        name = '_'.join(parts[-3:])
    
    # Max 40 chars
    if len(name) > 40:
        name = name[:40].rstrip('_')
    
    return name

def extract_base_pattern(filename):
    """Remove dates/regions from filename to find base pattern."""
    base = filename.lower()
    
    # Remove month patterns (Month_YY)
    base = re.sub(r'_(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)_\d{2,4}', '', base)
    # Remove year patterns
    base = re.sub(r'_\d{4}', '', base)
    
    # sanitize_name will do additional date stripping
    return sanitize_name(base)

def process_resources(resources, source_url, enable_preview=False):
    """Process each resource by type, return grouped files."""
    processed = []
    
    for res in resources:
        url = res['url']
        file_type = res['type']
        
        print(f"  Processing {file_type}: {Path(urlparse(url).path).name}", file=sys.stderr)
        
        if file_type == 'ZIP':
            files_inside = inspect_zip(url)
            print(f"    → {len(files_inside)} files inside", file=sys.stderr)
            
            for filename in files_inside:
                processed.append({
                    'url': url,
                    'extract': filename,
                    'period': extract_period_from_filename(filename) or extract_period_from_filename(source_url),  # Fallback to source URL
                    'pattern': extract_base_pattern(filename),
                    'original_type': 'ZIP'
                })
        
        elif file_type in ['XLSX', 'XLS', 'XLSM']:
            sheets = get_excel_sheets(url)
            print(f"    → {len(sheets)} sheets: {', '.join(sheets[:3])}{'...' if len(sheets) > 3 else ''}", file=sys.stderr)
            
            # Use short base name from URL (e.g., "gp_appt" not full filename)
            base = sanitize_name(url)
            # Extract meaningful part (usually last 2-3 segments)
            parts = [p for p in base.split('_') if p and len(p) > 2]
            short_base = '_'.join(parts[-2:]) if len(parts) >= 2 else base
            
            for sheet in sheets:
                sheet_code = sanitize_name(sheet)
                pattern = f"{short_base}_{sheet_code}"
                
                processed.append({
                    'url': url,
                    'sheet': sheet,
                    'period': extract_period_from_filename(url) or extract_period_from_filename(source_url),  # Fallback to source URL
                    'pattern': pattern,
                    'original_type': file_type
                })
        
        elif file_type == 'CSV':
            processed.append({
                'url': url,
                'period': extract_period_from_filename(url) or extract_period_from_filename(source_url),  # Fallback to source URL
                'pattern': sanitize_name(url),
                'original_type': 'CSV'
            })

             
    # Second Pass for Previews (to keep loop clean)
    if enable_preview:
        print("  🔍 Generating content previews (Deep Dive)...", file=sys.stderr)
        from datawarp.utils.zip_handler import extract_file_from_zip
        from collections import defaultdict
        
        # Cache downloaded temp files by URL to avoid redundant network requests
        download_cache = {}
        
        # Group items by URL to detect multi-sheet files
        items_by_url = defaultdict(list)
        for item in processed:
            items_by_url[item['url']].append(item)
        
        try:
            # Process each URL's items
            for url, url_items in items_by_url.items():
                # Download once per URL
                print(f"    Peek: {Path(url).name}", file=sys.stderr)
                
                suffix = Path(url).suffix
                with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                    with requests.get(url, stream=True) as r:
                        r.raise_for_status()
                        for chunk in r.iter_content(chunk_size=8192):
                            tmp.write(chunk)
                    tmp_path = Path(tmp.name)
                    download_cache[url] = tmp_path
                
                # Detect if this is a multi-sheet Excel file
                excel_sheets = [item for item in url_items if item.get('sheet') and suffix.lower() in ['.xlsx', '.xls', '.xlsm']]
                use_batch_mode = len(excel_sheets) > 1
                
                batch_gen = None
                if use_batch_mode:
                    # BATCH MODE: Use optimized FileExtractor (row-major access)
                    batch_gen = BatchPreviewGenerator(tmp_path)
                
                # Generate previews for all items from this URL
                for item in url_items:
                    try:
                        target_path = tmp_path
                        
                        # Handle ZIP extraction if needed
                        if item.get('extract'):
                            try:
                                target_path = extract_file_from_zip(tmp_path, item['extract'])
                            except Exception as e:
                                print(f"      ⚠ Extract failed: {e}", file=sys.stderr)
                                target_path = None
                        
                        # Generate preview
                        if target_path:
                            if batch_gen and item.get('sheet'):
                                # Use optimized batch generator for Excel sheets
                                item['preview'] = batch_gen.generate_preview(item['sheet'])
                            else:
                                # Use lightweight preview for CSV or single sheets
                                item['preview'] = generate_preview(target_path, item.get('sheet'))
                        
                        # Cleanup extracted ZIP member
                        try:
                            if target_path and target_path != tmp_path and target_path.exists():
                                target_path.unlink()
                                if item.get('extract'):
                                    target_path.parent.rmdir()
                        except Exception:
                            pass
                    
                    except Exception as e:
                        print(f"      ⚠ Preview failed: {e}", file=sys.stderr)
                
                # Close batch generator if used
                if batch_gen:
                    batch_gen.close()
        
        finally:
            # Cleanup cached downloads
            for path in download_cache.values():
                try:
                    if path.exists():
                        path.unlink()
                except Exception:
                    pass
                 
                 
    return processed

class BatchPreviewGenerator:
    """Optimized preview generation for multiple sheets from same workbook.
    
    Loads workbook ONCE, generates previews for all sheets.
    Uses FileExtractor's structure inference logic but with cached workbook.
    
    ONLY used for multi-sheet files (2+ sheets). Single sheets use regular path.
    """
    
    def __init__(self, filepath: Path):
        """Load workbook once and cache it."""
        import openpyxl
        self.filepath = filepath
        self.wb = openpyxl.load_workbook(str(filepath), data_only=True)
        print(f"    [BATCH] Loaded workbook once: {filepath.name}", file=sys.stderr)
    
    def generate_preview(self, sheet_name: str) -> str:
        """Generate preview for a specific sheet using cached workbook.
        
        Uses FileExtractor with preview_mode=True for speed.
        """
        try:
            from datawarp.core.extractor import FileExtractor
            
            # DEBUG: Show which sheet we're processing
            print(f"      [BATCH] Processing sheet: {sheet_name}", file=sys.stderr)
            
            # Create FileExtractor with our cached workbook AND preview_mode
            extractor = FileExtractor(
                str(self.filepath), 
                sheet_name=sheet_name, 
                workbook=self.wb,
                preview_mode=True  # Skip merged cell processing (10-20x faster)
            )
            
            print(f"      [BATCH] Calling infer_structure()...", file=sys.stderr)
            struct = extractor.infer_structure()

            # Log sheet classification
            sheet_type_str = struct.sheet_type.name if struct.sheet_type else "UNKNOWN"
            if struct.sheet_type.name == "TABULAR":
                print(f"      [BATCH] [OK] {sheet_name}: {sheet_type_str} ({struct.total_data_rows} rows, {struct.total_data_cols} cols)", file=sys.stderr)
            elif struct.sheet_type.name == "METADATA":
                print(f"      [BATCH] [SKIP] {sheet_name}: {sheet_type_str} (metadata/contents sheet)", file=sys.stderr)
            elif struct.sheet_type.name == "EMPTY":
                print(f"      [BATCH] [SKIP] {sheet_name}: {sheet_type_str} (empty sheet)", file=sys.stderr)
            else:
                print(f"      [BATCH] [WARN] {sheet_name}: {sheet_type_str}", file=sys.stderr)
            
            # Format preview using FileExtractor's structured output
            sorted_cols = sorted(struct.columns.values(), key=lambda c: c.col_index)
            
            headers = []
            for c in sorted_cols:
                full_header = ' '.join(h for h in c.original_headers if h)
                quoted_header = f'"{full_header}"' if full_header else f'"{c.pg_name}"'
                headers.append(quoted_header)
            
            preview_lines = [", ".join(headers)]
            
            num_samples = len(sorted_cols[0].sample_values) if sorted_cols else 0
            for i in range(min(3, num_samples)):
                row_vals = []
                for col in sorted_cols:
                    if i < len(col.sample_values):
                        val = str(col.sample_values[i]).replace('\n', ' ').strip()
                        if len(val) > 50: 
                            val = val[:47] + "..."
                        row_vals.append(val)
                    else:
                        row_vals.append("")
                preview_lines.append(", ".join(row_vals))
            
            return "\n".join(preview_lines)
            
        except Exception as e:
            return f"Preview unavailable: {e}"
    
    def close(self):
        """Close the cached workbook."""
        try:
            self.wb.close()
        except:
            pass

def group_files(processed):
    """Group files by pattern."""
    groups = defaultdict(list)
    for item in processed:
        groups[item['pattern']].append(item)
    return groups

def generate_preview(filepath, sheet_name=None):
    """Generate content preview FAST using raw openpyxl (bypass FileExtractor).
    
    This is 10-20x faster than FileExtractor.infer_structure() because we:
    - Skip header detection, merged cell processing, type inference 
    - Just grab first 50 rows and format them
    """
    try:
        import openpyxl
        from itertools import islice
        
        path = Path(filepath)
        ext = path.suffix.lower()
        
        if ext == '.csv':
            # CSV: Use pandas for quick preview
            import pandas as pd
            try:
                df = pd.read_csv(filepath, nrows=10)
                preview_lines = [', '.join(f'"{col}"' for col in df.columns)]
                for _, row in df.head(3).iterrows():
                    preview_lines.append(', '.join(str(val)[:50] for val in row.values))
                return '\n'.join(preview_lines)
            except Exception as e:
                return f"Preview unavailable: {e}"
                
        elif ext in ['.xlsx', '.xls', '.xlsm']:
            # Excel: Use raw openpyxl (FAST - no structure inference)
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return f"Preview unavailable: sheet '{sheet_name}' not found"
                ws = wb[sheet_name]
            else:
                ws = wb.worksheets[0]
            
            # Read first 50 rows (much faster than full structure inference)
            preview_lines = []
            for row in islice(ws.iter_rows(values_only=True), 50):
                if any(row):  # Skip empty rows
                    # Format row: truncate long values, join with commas
                    formatted = ', '.join(
                        str(val)[:50].replace('\n', ' ').strip() if val else ''
                        for val in row[:15]  # First 15 columns
                    )
                    if formatted.strip():
                        preview_lines.append(formatted)
                
                # Stop after we have 10 non-empty lines
                if len(preview_lines) >= 10:
                    break
            
            wb.close()
            return '\n'.join(preview_lines) if preview_lines else "Preview unavailable: empty sheet"
        else:
            return f"Preview unavailable: unsupported format {ext}"
            
    except Exception as e:
        return f"Preview unavailable: {e}"

def generate_sources(groups):
    """Create source configs from grouped files."""
    sources = []
    
    for pattern, files in groups.items():
        # Auto-generate names
        code = pattern
        name = pattern.replace('_', ' ').title()
        table = f"tbl_{pattern}"
        
        # Sort files by period
        files.sort(key=lambda f: f['period'] if f['period'] else 'zzzz')
        
        # Build file entries
        file_entries = []
        for f in files:
            entry = {
                'url': f['url'],
                'period': f['period'],
                'mode': 'append' if f.get('extract') else 'replace'
            }
            
            if f.get('extract'):
                entry['extract'] = f['extract']
            if f.get('sheet'):
                entry['sheet'] = f['sheet']
            if f.get('preview'):
                entry['preview'] = f['preview']
            
            file_entries.append(entry)
        
        source = {
            'code': code,
            'name': name,
            'table': table,
            'enabled': True,  # User can set to False to skip loading
            'files': file_entries
        }
        
        sources.append(source)
    
    return sources

def main():
    import argparse
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument('url', help='Publication URL')
    parser.add_argument('output', help='Output manifest file')
    parser.add_argument('--skip-preview', action='store_true', 
                       help='Skip downloading files for content previews (NOT RECOMMENDED - LLM needs previews to avoid hallucinations)')
    
    args = parser.parse_args()
    
    pub_url = args.url
    output_file = args.output
    enable_preview = not args.skip_preview  # Default: previews enabled

    
    # Scrape
    resources = scrape_resources(pub_url)
    if not resources:
        print("No files found!", file=sys.stderr)
        sys.exit(1)
    
    # Capture context (optional, for LLM enrichment)
    pub_context = capture_publication_context(pub_url, resources)
    
    # Process
    processed = process_resources(resources, pub_url, enable_preview)  # Pass source URL for fallback
    
    # Group
    groups = group_files(processed)
    
    # Generate
    sources = generate_sources(groups)
    
    # Create manifest
    pub_name = Path(urlparse(pub_url).path).name.replace('-', '_')
    manifest = {
        'manifest': {
            'name': f"{pub_name}_{datetime.now().strftime('%Y%m%d')}",
            'description': f"Generated from {pub_url}",
            'created_at': datetime.now().strftime('%Y-%m-%d'),
            'source_url': pub_url
        },
        'sources': sources
    }
    
    # Add context if captured
    if pub_context:
        manifest['manifest']['publication_context'] = pub_context
    
    # Write with block style for multi-line strings
    def str_representer(dumper, data):
        """Use block literal style for strings with newlines."""
        if '\n' in data:
            return dumper.represent_scalar('tag:yaml.org,2002:str', data, style='|')
        return dumper.represent_scalar('tag:yaml.org,2002:str', data)
    
    yaml.add_representer(str, str_representer)
    yaml_output = yaml.dump(manifest, sort_keys=False, default_flow_style=False, allow_unicode=True)
    
    with open(output_file, 'w') as f:
        f.write(yaml_output)
    
    # Summary
    total_files = sum(len(s['files']) for s in sources)
    print(f"\n✅ Generated {len(sources)} sources with {total_files} files", file=sys.stderr)
    print(f"📝 Written to {output_file}", file=sys.stderr)
    print(f"👀 Review and edit to remove unwanted sheets/sources", file=sys.stderr)
    print(f"🚀 Load: datawarp load-batch {output_file}", file=sys.stderr)

if __name__ == '__main__':
    main()
