"""File download utilities for DataWarp v2."""

import tempfile
import logging
from pathlib import Path
import requests

logger = logging.getLogger(__name__)

# Session-scoped download cache: url → local path
# Prevents re-downloading same file for multiple sheets
_download_cache: dict[str, Path] = {}


def clear_download_cache():
    """Clear the download cache. Call at end of batch processing."""
    _download_cache.clear()


def download_file(url: str) -> Path:
    """Download file to temp location. Handles URLs and local paths.

    Args:
        url: URL or file path

    Returns:
        Path to file (local or downloaded temp file)

    Raises:
        requests.HTTPError: If download fails
        FileNotFoundError: If local file doesn't exist
    """
    # Handle local paths
    if url.startswith('file://'):
        return Path(url[7:])

    # Check if it's a local file path
    path = Path(url)
    if path.exists():
        return path

    # Check cache first - avoid re-downloading same file for multiple sheets
    if url in _download_cache:
        cached_path = _download_cache[url]
        if cached_path.exists():
            logger.debug(f"Using cached download: {cached_path.name}")
            return cached_path

    # Download without progress bar (batch.py shows inline progress)
    from urllib.parse import urlparse

    response = requests.get(url, stream=True, timeout=60)
    response.raise_for_status()
    
    # Determine file extension from URL
    suffix = Path(urlparse(url).path).suffix or '.xlsx'
    
    # Simple download
    temp_file = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    
    for chunk in response.iter_content(chunk_size=8192):
        temp_file.write(chunk)
    
    temp_file.close()
    
    filepath = Path(temp_file.name)

    # Auto-convert old .xls files to .xlsx format
    # FileExtractor only supports .xlsx (via openpyxl)
    if filepath.suffix.lower() == '.xls':
        logger.info(f"Converting .xls to .xlsx: {filepath.name}")
        filepath = convert_xls_to_xlsx(filepath)

    # Cache the downloaded file path
    _download_cache[url] = filepath
    logger.debug(f"Downloaded and cached: {filepath.name}")

    return filepath


def convert_xls_to_xlsx(xls_path: Path) -> Path:
    """
    Convert old Excel .xls format to modern .xlsx format.

    This enables FileExtractor (which uses openpyxl) to process
    legacy Excel files that use the old binary format.

    Args:
        xls_path: Path to .xls file

    Returns:
        Path to converted .xlsx file

    Raises:
        ImportError: If xlrd not installed
        Exception: If conversion fails
    """
    try:
        import xlrd
        from openpyxl import Workbook
    except ImportError as e:
        raise ImportError(
            "xlrd library required for .xls file support. "
            "Install with: pip install xlrd"
        ) from e

    logger.debug(f"Reading .xls file: {xls_path}")

    # Read the old .xls file
    xls_book = xlrd.open_workbook(str(xls_path), formatting_info=False)

    # Create new .xlsx file path (same name, different extension)
    xlsx_path = xls_path.with_suffix('.xlsx')

    # Create new workbook
    xlsx_book = Workbook()
    xlsx_book.remove(xlsx_book.active)  # Remove default sheet

    # Copy each sheet
    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

        # Copy all cells
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                cell_type = xls_sheet.cell_type(row_idx, col_idx)

                # xlrd cell types: 0=empty, 1=text, 2=number, 3=date, 4=boolean, 5=error
                # Convert to appropriate Python type
                if cell_type == 0:  # Empty
                    value = None
                elif cell_type == 1:  # Text
                    value = cell_value
                elif cell_type == 2:  # Number
                    value = cell_value
                elif cell_type == 3:  # Date
                    # xlrd returns dates as floats, convert to Excel date number
                    value = cell_value
                elif cell_type == 4:  # Boolean
                    value = bool(cell_value)
                elif cell_type == 5:  # Error
                    value = None
                else:
                    value = cell_value

                # Write to xlsx (openpyxl uses 1-based indexing)
                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=value)

    # Save the converted file
    xlsx_book.save(str(xlsx_path))
    logger.info(f"Converted {xls_path.name} → {xlsx_path.name}")

    return xlsx_path
