"""
Manifest generation from NHS publication URLs.

Extracted from scripts/url_to_manifest.py for library use.
"""

import re
import tempfile
import yaml
import requests
from pathlib import Path
from datetime import datetime
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
from collections import defaultdict
from zipfile import ZipFile
from dataclasses import dataclass
from typing import Optional

from datawarp.supervisor.events import EventStore, EventType, EventLevel, create_event


MONTHS = {'jan':1,'feb':2,'mar':3,'apr':4,'may':5,'jun':6,
          'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12}

MONTHS_FULL = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
               'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}


# === INTELLIGENT ADAPTIVE SAMPLING ===

def _detect_column_pattern(columns: list) -> dict:
    """Detect repetitive sequential patterns in column names.

    Returns:
        Pattern info dict with has_pattern, pattern_prefix, pattern_count, unique_count
    """
    from collections import defaultdict

    pattern_groups = defaultdict(list)

    for col in columns:
        # Extract prefix before numbers: "week_52_53" -> "week"
        match = re.match(r'^([a-z_]+?)_*(\d+)', col, re.IGNORECASE)
        if match:
            prefix = match.group(1)
            pattern_groups[prefix].append(col)

    # Pattern detected if 10+ columns share prefix with numbers
    for prefix, cols in pattern_groups.items():
        if len(cols) >= 10:
            return {
                'has_pattern': True,
                'pattern_prefix': prefix,
                'pattern_cols': cols,
                'pattern_count': len(cols),
                'unique_count': len(columns) - len(cols)
            }

    return {
        'has_pattern': False,
        'pattern_count': 0,
        'unique_count': len(columns)
    }


def _adaptive_sample_rows(columns: list, sample_rows: list) -> tuple:
    """Intelligently sample sample_rows based on column patterns and count.

    Strategy:
    - Files ≤ 50 columns: Full sampling
    - Pattern files (10+ sequential): Sample 3 pattern + all unique
    - Non-pattern ≤ 75 columns: Full sampling (conservative)
    - Non-pattern > 75 columns: Stratified sampling (30 columns)

    Returns:
        (filtered_sample_rows, sampling_info)
    """
    col_count = len(columns)

    # Threshold 1: Small files - no sampling needed
    if col_count <= 50:
        return sample_rows, {'strategy': 'full', 'sampled': col_count}

    # Detect patterns
    pattern_info = _detect_column_pattern(columns)

    # Phase 1: Pattern-aware sampling
    if pattern_info['has_pattern']:
        pattern_cols = pattern_info['pattern_cols']
        unique_cols = [c for c in columns if c not in pattern_cols]

        # Sample pattern: first, middle, last
        pattern_samples = [
            pattern_cols[0],
            pattern_cols[len(pattern_cols)//2],
            pattern_cols[-1]
        ] if len(pattern_cols) >= 3 else pattern_cols

        # Keep ALL unique columns + pattern samples
        sample_cols = unique_cols + pattern_samples

        # Filter sample_rows
        filtered_rows = [
            {col: row[col] for col in sample_cols if col in row}
            for row in sample_rows
        ]

        return filtered_rows, {
            'strategy': 'pattern_aware',
            'pattern_count': len(pattern_cols),
            'unique_count': len(unique_cols),
            'sampled': len(sample_cols),
            'coverage_unique': '100%'
        }

    # Threshold 2: Medium files without patterns - be conservative
    if col_count <= 75:
        return sample_rows, {
            'strategy': 'full',
            'reason': 'mostly_unique_below_threshold',
            'sampled': col_count
        }

    # Phase 2: Large unique files - stratified sampling
    sample_cols = _stratified_sample(columns, target=30)

    filtered_rows = [
        {col: row[col] for col in sample_cols if col in row}
        for row in sample_rows
    ]

    return filtered_rows, {
        'strategy': 'stratified',
        'unique_count': col_count,
        'sampled': len(sample_cols),
        'coverage': f'{100 * len(sample_cols) / col_count:.1f}%'
    }


def _stratified_sample(columns: list, target: int = 30) -> list:
    """Sample columns across different types for diversity.

    Categories: IDs, dates, names, measures, codes, other
    """
    categories = {
        'ids': [],
        'dates': [],
        'names': [],
        'measures': [],
        'codes': [],
        'other': []
    }

    # Categorize columns
    for col in columns:
        col_lower = col.lower()
        if ('_id' in col_lower or '_code' in col_lower) and 'diagnosis' not in col_lower:
            categories['ids'].append(col)
        elif '_date' in col_lower or '_time' in col_lower:
            categories['dates'].append(col)
        elif '_name' in col_lower or '_description' in col_lower:
            categories['names'].append(col)
        elif any(x in col_lower for x in ['_count', '_total', '_percentage', '_score', '_number']):
            categories['measures'].append(col)
        elif any(x in col_lower for x in ['diagnosis_', 'procedure_', 'medication_']):
            categories['codes'].append(col)
        else:
            categories['other'].append(col)

    # Sample proportionally
    sampled = []

    # Always keep all IDs (important context)
    sampled.extend(categories['ids'])
    remaining = target - len(sampled)

    # Distribute remaining across other categories
    for category in ['dates', 'names', 'measures', 'codes', 'other']:
        items = categories[category]
        if items and remaining > 0:
            # Take evenly distributed samples
            if len(items) >= 3:
                sampled.extend([items[0], items[len(items)//2], items[-1]])
            elif len(items) == 2:
                sampled.extend(items)
            elif len(items) == 1:
                sampled.extend(items)
            remaining = target - len(sampled)

    return sampled[:target]


@dataclass
class ManifestResult:
    """Result of manifest generation."""
    success: bool
    manifest_path: Optional[Path] = None
    sources_count: int = 0
    files_count: int = 0
    error: Optional[str] = None


def parse_period(text):
    """Extract period from text (filenames, URLs)."""
    text = text.lower()

    # Pattern 1: Full month-YYYY (in URL paths): 31-october-2025, october-2025
    for month, num in MONTHS_FULL.items():
        match = re.search(rf'\d*-?{month}-(\d{{4}})', text)
        if match:
            return f"{match.group(1)}-{num:02d}"

    # Pattern 2: Month-YYYY (4-digit year): sep-2024 → 2024-09
    match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-_](\d{4})', text)
    if match:
        return f"{match.group(2)}-{MONTHS[match.group(1)]:02d}"

    # Pattern 3: Month_YY (2-digit year): apr_25 → 2025-04
    match = re.search(r'(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[-_](\d{2})(?:[^0-9]|$)', text)
    if match:
        return f"20{match.group(2)}-{MONTHS[match.group(1)]:02d}"

    # Pattern 4: Year only: 2023 → 2023-01
    match = re.search(r'[/_](\d{4})(?:[/_]|$)', text)
    if match:
        return f"{match.group(1)}-01"

    return None


def extract_period_from_filename(filename: str) -> str | None:
    """Extract period from filename patterns."""
    # Extract just the filename from URL/path (ignore directory dates like /2026/01/)
    from pathlib import Path
    from urllib.parse import urlparse
    if filename.startswith('http'):
        filename = Path(urlparse(filename).path).name
    else:
        filename = Path(filename).name
    filename_lower = filename.lower()

    # Quarterly: q1-2526, q2-2425, q3-2324
    match = re.search(r'q([1-4])[_-]?(\d{2})(\d{2})', filename_lower)
    if match:
        quarter, year1, year2 = match.groups()
        year = f"20{year1}"
        return f"{year}-Q{quarter}"

    # Monthly patterns
    month_map = {
        'jan': '01', 'feb': '02', 'mar': '03', 'apr': '04',
        'may': '05', 'jun': '06', 'jul': '07', 'aug': '08',
        'sep': '09', 'oct': '10', 'nov': '11', 'dec': '12'
    }

    for month_name, month_num in month_map.items():
        if month_name in filename_lower:
            # Try 4-digit year first (e.g., nov2025)
            year_match = re.search(r'20\d{2}', filename_lower)
            if year_match:
                return f"{year_match.group()}-{month_num}"
            # Try 2-digit year (e.g., nov25 -> 2025-11)
            two_digit_match = re.search(rf'{month_name}[_-]?(\d{{2}})(?!\d)', filename_lower)
            if two_digit_match:
                year = f"20{two_digit_match.group(1)}"
                return f"{year}-{month_num}"

    # Numeric month: 2025-10, m10-2025
    match = re.search(r'(20\d{2})[_-](\d{1,2})', filename_lower)
    if match:
        year, month = match.groups()
        return f"{year}-{month.zfill(2)}"

    # Weekly: week42-2025, w42-2025
    match = re.search(r'w(?:eek)?[_-]?(\d{1,2})[_-]?(20\d{2})', filename_lower)
    if match:
        week, year = match.groups()
        return f"{year}-W{week.zfill(2)}"

    # Annual: 2025, fy2025
    match = re.search(r'(?:fy)?[_-]?(20\d{2})(?![_-]\d)', filename_lower)
    if match and not re.search(r'(q[1-4]|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)', filename_lower):
        return match.group(1)

    return None


def sanitize_name(text):
    """Create clean, SHORT, DATE-AGNOSTIC code from filename."""
    name = Path(text).stem
    name = re.sub(r'%20', '_', name)
    name = re.sub(r'[^a-z0-9_]', '_', name.lower())
    name = re.sub(r'_+', '_', name).strip('_')

    # Strip ALL date patterns for stable table names
    name = re.sub(r'_(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)_\d{2,4}', '', name, flags=re.I)
    name = re.sub(r'_(january|february|march|april|may|june|july|august|september|october|november|december)_\d{2,4}', '', name, flags=re.I)
    name = re.sub(r'_\d{4}(?:_|$)', '_', name)
    name = re.sub(r'_(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)(?:_|$)', '_', name, flags=re.I)

    name = re.sub(r'_+', '_', name).strip('_')

    # Keep only last 2-3 meaningful segments
    parts = [p for p in name.split('_') if p and len(p) > 2]
    if len(parts) > 3:
        name = '_'.join(parts[-3:])

    # Max 40 chars
    if len(name) > 40:
        name = name[:40].rstrip('_')

    return name


def extract_base_pattern(filename):
    """Remove dates/regions from filename to find base pattern."""
    base = filename.lower()
    base = re.sub(r'_(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)_\d{2,4}', '', base)
    base = re.sub(r'_\d{4}', '', base)
    return sanitize_name(base)


def scrape_resources(url, event_store: EventStore = None):
    """Scrape ALL downloadable resources (ZIP/XLSX/CSV)."""
    if event_store:
        event_store.emit(create_event(
            EventType.STAGE_STARTED,
            event_store.run_id,
            message=f"Scraping {url}",
            stage="scrape",
            level=EventLevel.DEBUG
        ))

    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    }

    response = requests.get(url, headers=headers)
    response.raise_for_status()

    soup = BeautifulSoup(response.content, 'html.parser')
    resources = []

    for link in soup.find_all('a', href=True):
        href = link['href']
        lower_href = href.lower()

        if lower_href.endswith(('.zip', '.xlsx', '.xls', '.xlsm', '.csv')):
            full_url = urljoin(url, href)
            file_type = lower_href.split('.')[-1].upper()
            title = link.get_text(strip=True) or Path(urlparse(full_url).path).name

            resources.append({
                'url': full_url,
                'type': file_type,
                'title': title
            })

    if event_store:
        event_store.emit(create_event(
            EventType.STAGE_COMPLETED,
            event_store.run_id,
            message=f"Found {len(resources)} files",
            stage="scrape",
            files_found=len(resources)
        ))

    return resources


def get_excel_sheets(url):
    """Download XLSX and return sheet names."""
    import openpyxl
    from io import BytesIO

    try:
        response = requests.get(url)
        wb = openpyxl.load_workbook(BytesIO(response.content), read_only=True, data_only=True)
        return wb.sheetnames
    except Exception:
        return []


def inspect_zip(url):
    """Download ZIP and list CSV/XLSX files."""
    response = requests.get(url, stream=True)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
    for chunk in response.iter_content(8192):
        tmp.write(chunk)
    tmp.close()

    with ZipFile(tmp.name) as zf:
        files = [f for f in zf.namelist()
                if not f.endswith('/') and f.lower().endswith(('.csv', '.xlsx', '.xls'))]

    Path(tmp.name).unlink()
    return files


def is_metadata_sheet(sheet_name: str) -> bool:
    """Detect if sheet is metadata/contents/notes based on name patterns."""
    sheet_lower = sheet_name.lower()

    metadata_patterns = [
        'title', 'cover', 'contents', 'notes', 'metadata', 'glossary',
        'definitions', 'data quality', 'guidance', 'readme', 'about',
        'description', 'column', 'metric', 'methodology', 'contact'
    ]

    return any(pattern in sheet_lower for pattern in metadata_patterns)


def process_resources(resources, source_url, event_store: EventStore = None):
    """Process each resource by type, return grouped files."""
    processed = []

    for res in resources:
        url = res['url']
        file_type = res['type']

        if event_store:
            event_store.emit(create_event(
                EventType.STAGE_STARTED,
                event_store.run_id,
                message=f"Processing {file_type}: {Path(urlparse(url).path).name}",
                stage="process",
                level=EventLevel.DEBUG
            ))

        if file_type == 'ZIP':
            files_inside = inspect_zip(url)
            for filename in files_inside:
                # Emit event for each file in ZIP (like we do for sheets in XLSX)
                if event_store:
                    file_ext = Path(filename).suffix.upper()[1:]  # .csv -> CSV
                    event_store.emit(create_event(
                        EventType.SHEET_CLASSIFIED,
                        event_store.run_id,
                        message=f"File in ZIP: {filename} ({file_ext})",
                        stage="process",
                        sheet_name=filename,  # Use same field for consistency
                        sheet_type=file_ext,   # CSV, XLSX, etc.
                        level=EventLevel.DEBUG
                    ))

                processed.append({
                    'url': url,
                    'extract': filename,
                    'period': extract_period_from_filename(filename) or extract_period_from_filename(source_url),
                    'pattern': extract_base_pattern(filename),
                    'original_type': 'ZIP',
                    'enabled': True  # CSV files in ZIPs are always data
                })

        elif file_type in ['XLSX', 'XLS', 'XLSM']:
            sheets = get_excel_sheets(url)
            base = sanitize_name(url)
            parts = [p for p in base.split('_') if p and len(p) > 2]
            short_base = '_'.join(parts[-2:]) if len(parts) >= 2 else base

            for sheet in sheets:
                sheet_code = sanitize_name(sheet)
                pattern = f"{short_base}_{sheet_code}"

                # Classify sheet
                is_metadata = is_metadata_sheet(sheet)

                if event_store:
                    sheet_type = "METADATA" if is_metadata else "TABULAR"
                    event_store.emit(create_event(
                        EventType.SHEET_CLASSIFIED,
                        event_store.run_id,
                        message=f"Sheet: {sheet} ({sheet_type})",
                        stage="process",
                        sheet_name=sheet,
                        sheet_type=sheet_type,
                        level=EventLevel.DEBUG
                    ))

                processed.append({
                    'url': url,
                    'sheet': sheet,
                    'period': extract_period_from_filename(url) or extract_period_from_filename(source_url),
                    'pattern': pattern,
                    'original_type': file_type,
                    'enabled': not is_metadata  # Disable metadata sheets
                })

        elif file_type == 'CSV':
            processed.append({
                'url': url,
                'period': extract_period_from_filename(url) or extract_period_from_filename(source_url),
                'pattern': sanitize_name(url),
                'original_type': 'CSV',
                'enabled': True  # CSVs are always data
            })

    return processed


def group_files(processed):
    """Group files by pattern."""
    groups = defaultdict(list)
    for item in processed:
        groups[item['pattern']].append(item)
    return groups


def generate_sources(groups):
    """Create source configs from grouped files."""
    sources = []

    for pattern, files in groups.items():
        code = pattern
        name = pattern.replace('_', ' ').title()
        table = f"tbl_{pattern}"

        files.sort(key=lambda f: f['period'] if f['period'] else 'zzzz')

        # Check if any file in group is enabled (data sheet)
        is_enabled = any(f.get('enabled', True) for f in files)

        file_entries = []
        for f in files:
            entry = {
                'url': f['url'],
                'period': f['period'],
                'mode': 'append' if f.get('extract') else 'replace'
            }

            if f.get('extract'):
                entry['extract'] = f['extract']
            if f.get('sheet'):
                entry['sheet'] = f['sheet']

            file_entries.append(entry)

        source = {
            'code': code,
            'name': name,
            'table': table,
            'enabled': is_enabled,  # Use classification result
            'files': file_entries
        }

        sources.append(source)

    return sources


def add_file_preview(file_entry: dict, event_store: EventStore = None) -> dict:
    """Download file and add column preview for LLM enrichment.

    Uses FileExtractor for Excel files to properly detect headers in files with metadata sections.
    Returns file_entry with 'preview' field added containing actual column names.
    """
    import pandas as pd
    from pathlib import Path
    from datawarp.utils.download import download_file

    url = file_entry['url']
    file_name = Path(urlparse(url).path).name
    file_ext = Path(file_name).suffix.lower()

    try:
        # Download file using cached download utility (avoids re-downloading for multiple sheets)
        tmp_path = download_file(url)

        try:
            if file_ext == '.csv':
                # CSV: Simple pandas read (CSVs don't have metadata sections)
                df = pd.read_csv(tmp_path, nrows=3)
                columns = df.columns.tolist()
                sample_rows = df.head(3).to_dict('records')

                # Intelligent adaptive sampling for large CSV files
                sample_rows, sampling_info = _adaptive_sample_rows(columns, sample_rows)

                file_entry['preview'] = {
                    'columns': columns,
                    'sample_rows': sample_rows
                }

                if sampling_info['strategy'] != 'full' and event_store:
                    event_store.emit(create_event(
                        EventType.STAGE_COMPLETED,
                        event_store.run_id,
                        message=f"Adaptive sampling: {sampling_info['strategy']} ({sampling_info.get('sampled', len(columns))} of {len(columns)} columns)",
                        stage="preview",
                        level=EventLevel.DEBUG,
                        context=sampling_info
                    ))

            elif file_ext in ['.xlsx', '.xls', '.xlsm']:
                # Excel: Use FileExtractor to correctly detect headers (handles metadata sections)
                from datawarp.core.extractor import FileExtractor
                import openpyxl

                sheet_name = file_entry.get('sheet', 0)

                # Use FileExtractor to detect structure (pass filepath, not worksheet)
                extractor = FileExtractor(str(tmp_path), sheet_name=sheet_name)
                structure = extractor.infer_structure()

                if structure.is_valid:
                    # Get actual column names from detected structure
                    columns = [col.pg_name for col in structure.columns.values()]

                    # Sample first 3 data rows (not metadata rows)
                    sample_rows = []
                    data_start = structure.data_start_row
                    data_end = min(data_start + 2, structure.data_end_row)  # Up to 3 rows

                    # Load workbook to read sample rows
                    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
                    ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

                    for row_num in range(data_start, data_end + 1):
                        row_dict = {}
                        for col_idx, col_info in structure.columns.items():
                            cell_value = ws.cell(row=row_num, column=col_idx).value
                            row_dict[col_info.pg_name] = cell_value
                        sample_rows.append(row_dict)

                    wb.close()

                    # Intelligent adaptive sampling for large files
                    sample_rows, sampling_info = _adaptive_sample_rows(columns, sample_rows)

                    file_entry['preview'] = {
                        'columns': columns,
                        'sample_rows': sample_rows
                    }

                    if sampling_info['strategy'] != 'full' and event_store:
                        event_store.emit(create_event(
                            EventType.STAGE_COMPLETED,
                            event_store.run_id,
                            message=f"Adaptive sampling: {sampling_info['strategy']} ({sampling_info.get('sampled', len(columns))} of {len(columns)} columns)",
                            stage="preview",
                            level=EventLevel.DEBUG,
                            context=sampling_info
                        ))
                else:
                    # Fallback to pandas if FileExtractor fails
                    df = pd.read_excel(tmp_path, sheet_name=sheet_name, nrows=3)
                    columns = df.columns.tolist()
                    sample_rows = df.head(3).to_dict('records')

                    file_entry['preview'] = {
                        'columns': columns,
                        'sample_rows': sample_rows
                    }

            if event_store and 'preview' in file_entry:
                event_store.emit(create_event(
                    EventType.STAGE_COMPLETED,
                    event_store.run_id,
                    message=f"Preview generated: {len(file_entry['preview']['columns'])} columns",
                    stage="preview",
                    level=EventLevel.DEBUG,
                    context={'file': file_name, 'columns': file_entry['preview']['columns']}
                ))

        except Exception as inner_e:
            # Log error but don't fail - preview is optional
            if event_store:
                event_store.emit(create_event(
                    EventType.WARNING,
                    event_store.run_id,
                    message=f"Preview extraction failed: {str(inner_e)}",
                    stage="preview",
                    level=EventLevel.WARNING,
                    context={'file': file_name, 'error': str(inner_e)}
                ))
            # Note: Downloaded file will be cleaned up by clear_download_cache()

    except Exception as e:
        if event_store:
            event_store.emit(create_event(
                EventType.WARNING,
                event_store.run_id,
                message=f"Preview generation failed: {str(e)}",
                stage="preview",
                level=EventLevel.WARNING,
                context={'file': file_name, 'error': str(e)}
            ))
        # Note: If no event_store (CLI mode), errors are silently ignored to allow
        # manifest generation to continue. This is acceptable as preview is optional.

    return file_entry


def generate_manifest(
    url: str,
    output_path: Path,
    event_store: EventStore = None,
    skip_preview: bool = False
) -> ManifestResult:
    """
    Generate DataWarp manifest from NHS publication URL.

    Supports two URL types (auto-detected):
    - Landing page: HTML page with links to files (scraped for .xlsx/.csv/.zip links)
    - Direct file: URL ending in .xlsx/.xls/.csv/.zip (used directly, no scraping)

    Args:
        url: Publication landing page URL OR direct file URL
        output_path: Where to write manifest YAML
        event_store: Optional event store for observability
        skip_preview: Skip downloading files for previews (default: False, previews enabled)

    Returns:
        ManifestResult with success status and counts
    """
    try:
        if event_store:
            event_store.emit(create_event(
                EventType.STAGE_STARTED,
                event_store.run_id,
                message="Generating manifest",
                stage="manifest"
            ))

        # Detect URL type: direct file vs landing page
        url_lower = url.lower()
        if url_lower.endswith(('.xlsx', '.xls', '.xlsm', '.csv', '.zip')):
            # Direct file URL - skip HTML scraping
            file_ext = url.split('.')[-1].upper()
            file_name = Path(urlparse(url).path).name
            resources = [{
                'url': url,
                'type': file_ext,
                'title': file_name
            }]
            if event_store:
                event_store.emit(create_event(
                    EventType.STAGE_COMPLETED,
                    event_store.run_id,
                    message=f"Direct file URL detected: {file_name}",
                    stage="scrape",
                    files_found=1
                ))
        else:
            # Landing page URL - scrape for file links
            resources = scrape_resources(url, event_store)

        if not resources:
            return ManifestResult(
                success=False,
                error="No files found"
            )

        # Process
        processed = process_resources(resources, url, event_store)

        # Group
        groups = group_files(processed)

        # Generate
        sources = generate_sources(groups)

        # Add previews (download files and extract columns)
        if not skip_preview:
            if event_store:
                event_store.emit(create_event(
                    EventType.STAGE_STARTED,
                    event_store.run_id,
                    message="Generating file previews",
                    stage="preview"
                ))

            for source in sources:
                for file_entry in source.get('files', []):
                    add_file_preview(file_entry, event_store)

            # Clear caches after preview generation
            from datawarp.utils.download import clear_download_cache
            from datawarp.core.extractor import clear_workbook_cache
            clear_download_cache()
            clear_workbook_cache()

        # Create manifest
        pub_name = Path(urlparse(url).path).name.replace('-', '_')
        manifest = {
            'manifest': {
                'name': f"{pub_name}_{datetime.now().strftime('%Y%m%d')}",
                'description': f"Generated from {url}",
                'created_at': datetime.now().strftime('%Y-%m-%d'),
                'source_url': url
            },
            'sources': sources
        }

        # Write YAML
        yaml_output = yaml.dump(manifest, sort_keys=False, default_flow_style=False, allow_unicode=True)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(yaml_output)

        # Summary
        total_files = sum(len(s['files']) for s in sources)

        if event_store:
            event_store.emit(create_event(
                EventType.STAGE_COMPLETED,
                event_store.run_id,
                message=f"Generated {len(sources)} sources with {total_files} files",
                stage="manifest",
                sources_count=len(sources),
                files_count=total_files
            ))

        return ManifestResult(
            success=True,
            manifest_path=output_path,
            sources_count=len(sources),
            files_count=total_files
        )

    except Exception as e:
        if event_store:
            event_store.emit(create_event(
                EventType.ERROR,
                event_store.run_id,
                message=f"Manifest generation failed: {str(e)}",
                stage="manifest",
                level=EventLevel.ERROR,
                error=str(e)
            ))
        return ManifestResult(
            success=False,
            error=str(e)
        )
